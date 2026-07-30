# -*- coding: utf-8 -*-
"""
exportar_inventario.py — ATRIO/CORPUS · inventários separados e compatíveis.

Mantém a interface antiga:
    py exportar_inventario.py --base ../ --nao-abrir

Gera, em corpus/_inventario/:
    inventario.csv                  -> alias de compatibilidade: inventário documental
    inventario.xlsx                 -> consolidado com abas Documentos e Pseudonimos
    inventario_documentos.csv       -> inventário documental
    inventario_pseudonimos.csv      -> inventário do cofre/pseudonimização

Não altera o CORPUS, o painel, o OCR nem o CAPTURA.bat.
"""

import argparse
import csv
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

# Pasta-mãe resolvida a partir da localização do próprio script, para rodar
# sem .bat e sem --base. Os .bat continuam passando --base ../ e funcionam igual.
BASE_PADRAO = str(Path(__file__).resolve().parent.parent)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: openpyxl não instalado. Rode: py -m pip install openpyxl")
    sys.exit(1)


DOCUMENTO_HEADERS = [
    "status",
    "processo",
    "arquivo",
    "extensao",
    "tamanho_bytes",
    "modificado_em",
    "pasta_base",
    "caminho_relativo",
]

PSEUDONIMO_HEADERS = [
    "tipo",
    "chave",
    "token",
    "valor_real",
    "origem",
    "severidade",
    "visto_em",
    "campo",
    "valor",
    "caminho_json",
]


PASTAS_DOCUMENTAIS = [
    ("arquivado_pseudonimizado", "01_pseudonimizado"),
    ("original_preservado", "00_originais"),
    ("segredo", "_segredo"),
    ("revisar_segredo", "_revisar_segredo"),
    ("ocr_pendente", "_ocr_pendente"),
    ("ocr_falhou", "_ocr_falhou"),
    ("duplicado", "_duplicados"),
]


def safe_rel(path: Path, base: Path) -> str:
    try:
        return str(path.relative_to(base))
    except ValueError:
        return str(path)


def inferir_processo(arquivo: Path, pasta_raiz: Path) -> str:
    try:
        rel = arquivo.relative_to(pasta_raiz)
        if len(rel.parts) >= 2:
            return rel.parts[0]
    except ValueError:
        pass
    return ""


def inventariar_documentos(base: Path) -> List[Dict[str, Any]]:
    corpus = base / "corpus"
    linhas: List[Dict[str, Any]] = []

    for status, pasta_nome in PASTAS_DOCUMENTAIS:
        pasta = corpus / pasta_nome
        if not pasta.exists():
            continue

        for arquivo in sorted(pasta.rglob("*")):
            if not arquivo.is_file():
                continue

            stat = arquivo.stat()
            linhas.append({
                "status": status,
                "processo": inferir_processo(arquivo, pasta),
                "arquivo": arquivo.name,
                "extensao": arquivo.suffix.lower().lstrip("."),
                "tamanho_bytes": stat.st_size,
                "modificado_em": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                "pasta_base": pasta_nome,
                "caminho_relativo": safe_rel(arquivo, corpus),
            })

    return linhas


def normalizar_valor(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (dict, list)):
        return json.dumps(valor, ensure_ascii=False)
    return str(valor)


def parece_item_pseudonimo(obj: Any) -> bool:
    if not isinstance(obj, dict):
        return False
    chaves = {str(k).lower() for k in obj.keys()}
    marcadores = {
        "token", "valor_real", "valor", "origem", "origens", "severidade",
        "visto_em", "last_seen", "first_seen", "tipo"
    }
    return bool(chaves & marcadores)


def extrair_linha_pseudonimo(path: Tuple[str, ...], obj: Dict[str, Any]) -> Dict[str, Any]:
    def get_any(*nomes: str) -> str:
        for nome in nomes:
            if nome in obj:
                return normalizar_valor(obj.get(nome))
        # tentativa case-insensitive
        lower_map = {str(k).lower(): k for k in obj.keys()}
        for nome in nomes:
            key = lower_map.get(nome.lower())
            if key is not None:
                return normalizar_valor(obj.get(key))
        return ""

    tipo = get_any("tipo")
    if not tipo and path:
        tipo = path[0].split("::", 1)[0]

    chave = path[-1] if path else ""

    return {
        "tipo": tipo,
        "chave": chave,
        "token": get_any("token"),
        "valor_real": get_any("valor_real", "real", "original"),
        "origem": get_any("origem", "origens", "processo", "origem_cnj"),
        "severidade": get_any("severidade", "risco"),
        "visto_em": get_any("visto_em", "last_seen", "first_seen", "data"),
        "campo": "",
        "valor": "",
        "caminho_json": ".".join(path),
    }


def achatar_cofre(obj: Any, path: Tuple[str, ...] = ()) -> Iterable[Dict[str, Any]]:
    """Extrai linhas legíveis do cofre sem depender rigidamente do formato interno."""
    if parece_item_pseudonimo(obj):
        yield extrair_linha_pseudonimo(path, obj)  # linha principal do item

        # Também preserva campos extras como trilha auditável.
        for k, v in obj.items():
            if isinstance(v, (dict, list)):
                continue
            campo = str(k)
            if campo.lower() in {"tipo", "token", "valor_real", "real", "original", "origem", "origens", "processo", "origem_cnj", "severidade", "risco", "visto_em", "last_seen", "first_seen", "data"}:
                continue
            linha = {h: "" for h in PSEUDONIMO_HEADERS}
            linha.update({
                "tipo": path[0].split("::", 1)[0] if path else "",
                "chave": path[-1] if path else "",
                "campo": campo,
                "valor": normalizar_valor(v),
                "caminho_json": ".".join(path + (campo,)),
            })
            yield linha
        return

    if isinstance(obj, dict):
        for k, v in obj.items():
            yield from achatar_cofre(v, path + (str(k),))
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            yield from achatar_cofre(v, path + (str(i),))
    else:
        # Folha simples: preserva para auditoria, mas sem fingir que é pseudônimo estruturado.
        linha = {h: "" for h in PSEUDONIMO_HEADERS}
        linha.update({
            "tipo": path[0].split("::", 1)[0] if path else "",
            "chave": path[-1] if path else "",
            "campo": path[-1] if path else "",
            "valor": normalizar_valor(obj),
            "caminho_json": ".".join(path),
        })
        yield linha


def inventariar_pseudonimos(base: Path) -> List[Dict[str, Any]]:
    cofre_path = base / "corpus" / "_cofre" / "cofre.json"
    if not cofre_path.exists():
        return []

    try:
        with cofre_path.open("r", encoding="utf-8") as f:
            cofre = json.load(f)
    except UnicodeDecodeError:
        with cofre_path.open("r", encoding="utf-8-sig") as f:
            cofre = json.load(f)

    linhas = list(achatar_cofre(cofre))

    # Remove linhas totalmente vazias, preservando ordem.
    limpas = []
    vistos = set()
    for linha in linhas:
        normalizada = {h: normalizar_valor(linha.get(h, "")) for h in PSEUDONIMO_HEADERS}
        assinatura = tuple(normalizada.get(h, "") for h in PSEUDONIMO_HEADERS)
        if not any(assinatura):
            continue
        if assinatura in vistos:
            continue
        vistos.add(assinatura)
        limpas.append(normalizada)

    return limpas


def salvar_csv(path: Path, headers: List[str], linhas: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for linha in linhas:
            writer.writerow({h: linha.get(h, "") for h in headers})


def escrever_aba(ws, titulo: str, headers: List[str], linhas: List[Dict[str, Any]]) -> None:
    ws.title = titulo[:31]
    ws.append(headers)
    for linha in linhas:
        ws.append([linha.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor="000000")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            texto = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(texto))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)


def salvar_xlsx(path: Path, documentos: List[Dict[str, Any]], pseudonimos: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    ws_doc = wb.active
    escrever_aba(ws_doc, "Documentos", DOCUMENTO_HEADERS, documentos)
    ws_pseudo = wb.create_sheet("Pseudonimos")
    escrever_aba(ws_pseudo, "Pseudonimos", PSEUDONIMO_HEADERS, pseudonimos)
    wb.save(path)


def abrir_arquivo(path: Path) -> None:
    if sys.platform.startswith("win"):
        subprocess.Popen(["start", "", str(path)], shell=True)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", default=BASE_PADRAO,
                    help=f"pasta-mãe do ATRIO (padrão: {BASE_PADRAO})")
    ap.add_argument("--nao-abrir", action="store_true", help="não abrir o XLSX ao final")
    args = ap.parse_args()

    base = Path(args.base).resolve()
    inventario_dir = base / "corpus" / "_inventario"
    inventario_dir.mkdir(parents=True, exist_ok=True)

    documentos = inventariar_documentos(base)
    pseudonimos = inventariar_pseudonimos(base)

    documentos_csv = inventario_dir / "inventario_documentos.csv"
    pseudonimos_csv = inventario_dir / "inventario_pseudonimos.csv"
    compat_csv = inventario_dir / "inventario.csv"
    consolidado_xlsx = inventario_dir / "inventario.xlsx"

    salvar_csv(documentos_csv, DOCUMENTO_HEADERS, documentos)
    salvar_csv(pseudonimos_csv, PSEUDONIMO_HEADERS, pseudonimos)
    salvar_csv(compat_csv, DOCUMENTO_HEADERS, documentos)
    salvar_xlsx(consolidado_xlsx, documentos, pseudonimos)

    print(f"Inventário documental CSV gerado: {documentos_csv}")
    print(f"Inventário de pseudônimos CSV gerado: {pseudonimos_csv}")
    print(f"Inventário consolidado XLSX gerado: {consolidado_xlsx}")
    print(f"  documentos={len(documentos)}  pseudonimos={len(pseudonimos)}")

    if not args.nao_abrir:
        abrir_arquivo(consolidado_xlsx)


if __name__ == "__main__":
    main()
